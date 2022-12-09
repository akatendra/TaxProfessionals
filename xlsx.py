import os
import openpyxl as xl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import logging
import logging.config

# Set up logging
import logging
import logging.config

logging.config.fileConfig("logging.ini", disable_existing_loggers=False)
logger = logging.getLogger(__name__)


def create_xlsx_file(file_name):
    wb = xl.Workbook()
    ws = wb.active
    #  Write col headers
    col_headers = ['personal_link',
                    'title',
                    'image_url',
                    'description',
                    'email',
                    'phone',
                    'website',
                    'instagram',
                    'facebook',
                    'street_address',
                    'address_locality',
                    'address_region',
                    'postal_code'
                   ]

    # Put headers into xlsx-file
    ws.append(col_headers)
    # Set styles to headers
    col_header_style(wb, ws)
    # Adjust columns width according content
    xlsx_file_adjust_col_width(ws)
    # Save xlsx-file
    wb.save(file_name)
    logging.debug(f'File {file_name} has been created!')


def append_xlsx_file(data, file_name):
    # Check if file exist
    if not os.path.isfile(file_name):
        create_xlsx_file(file_name)

    # Open a xlsx for reading
    wb = xl.load_workbook(filename=file_name)
    # Get the current Active Sheet
    ws = wb.active
    # You can also select a particular sheet
    # based on sheet name
    # ws = wb.get_sheet_by_name("Sheet1")

    # Put data into xlsx-file
    row = list(data.values())
    # Put data into xlsx-file
    ws.append(row)
    # Set styles to headers

    xlsx_file_adjust_col_width(ws)
    # Save xlsx-file
    wb.save(file_name)
    logging.debug(f'Data saved into file {file_name}!')


def xlsx_file_adjust_col_width(work_sheet):
    dims = {}
    for row in work_sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        work_sheet.column_dimensions[col].width = value * 1.2


def hyperlink_style(file_name):
    # Open a xlsx for reading
    wb = xl.load_workbook(filename=file_name)
    # Get the current Active Sheet
    work_sheet = wb.active
    for cell in work_sheet['A2':'A600']:
        cell[0].style = "Hyperlink"
    for cell in work_sheet['C2':'C600']:
        cell[0].style = "Hyperlink"
    for cell in work_sheet['G2':'G600']:
        cell[0].style = "Hyperlink"
    for cell in work_sheet['H2':'H600']:
        cell[0].style = "Hyperlink"
    for cell in work_sheet['I2':'I600']:
        cell[0].style = "Hyperlink"
    # Save xlsx-file
    wb.save(file_name)
    logging.debug(f'Huperlinks styles saved into file {file_name}!')


def col_header_style(work_book, work_sheet):
    if 'rladies_col_header' not in work_book.named_styles:
        col_header_style = NamedStyle(name='rladies_col_header')
        col_header_style.font = Font(bold=True, size=12)
        col_header_style.alignment = Alignment(horizontal='center')
        col_header_style.border = Border(
            left=Side(border_style='hair', color='FF000000'),
            right=Side(border_style='hair', color='FF000000'),
            top=Side(border_style='hair', color='FF000000'),
            bottom=Side(border_style='hair', color='FF000000'))
        work_book.add_named_style(col_header_style)
        # Set NamedStyle to 1st row
        for cell in work_sheet['1:1']:
            cell.style = 'rladies_col_header'


if __name__ == '__main__':
    # create_xlsx_file('r_ladies_us.xlsx')
    hyperlink_style('r_ladies_us.xlsx')
